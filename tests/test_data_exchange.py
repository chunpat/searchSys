import json
import unittest
from io import BytesIO
from pathlib import Path

import openpyxl

from app.data_exchange import export_xlsx, parse_and_validate


ROOT = Path(__file__).resolve().parents[1]


class DataExchangeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.source = json.loads((ROOT / "data" / "master_data_source.json").read_text(encoding="utf-8"))

    def test_standard_xlsx_round_trip(self):
        workbook = export_xlsx(self.source)
        restored, counts, errors = parse_and_validate(workbook, "xlsx", self.source)
        self.assertEqual(errors, [])
        self.assertEqual(counts["报价明细"], 345)
        self.assertEqual(len(restored["sku_mappings"]), 668)

    def test_json_round_trip(self):
        content = json.dumps(self.source, ensure_ascii=False).encode("utf-8")
        restored, counts, errors = parse_and_validate(content, "json", self.source)
        self.assertEqual(errors, [])
        self.assertEqual(counts["供应商能力"], 234)
        self.assertEqual(len(restored["issues"]), 1091)

    def test_export_explains_and_highlights_processing_states(self):
        workbook = openpyxl.load_workbook(BytesIO(export_xlsx(self.source)))
        info = workbook["导入说明"]
        self.assertEqual(info["A1"].value, "报价数据交换与处理说明")
        self.assertIn("必须处理", info["A6"].value)
        self.assertNotEqual(info["A6"].fill.fgColor.rgb, info["A7"].fill.fgColor.rgb)
        self.assertIn("不要修改工作表名", info["A19"].value)
        self.assertEqual(info["A23"].value, "状态怎么修改")
        self.assertIn("下拉箭头", info["A34"].value)
        self.assertIn("只改状态不会自动补出价格", info["A35"].value)

        guide = workbook["规则填写示例"]
        self.assertEqual(guide["A1"].value, "报价、规则与预估填写示例")
        self.assertIn("当前直接报价和历史预估使用这里", guide["A4"].value)
        self.assertEqual(guide["A13"].value, "固定价")
        self.assertEqual(guide["A17"].value, "A3排版")
        self.assertEqual(guide["A18"].value, "3D打印")
        self.assertEqual(guide["A31"].value, "四、固定单价怎么计价")
        self.assertEqual(guide["B38"].value, "=MAX(B33*B35+B36,B37)")
        self.assertEqual(guide["B38"].data_type, "f")
        self.assertIn("不自动乘数量", guide["A40"].value)

        quote_sheet = workbook["报价明细"]
        headers = {cell.value: cell.column for cell in quote_sheet[1]}
        price_status_column = headers["价格解析状态"]
        fills = {}
        for row in quote_sheet.iter_rows(min_row=2):
            status_cell = row[price_status_column - 1]
            if status_cell.value in {"已结构化", "待结构化", "缺失"}:
                fills.setdefault(status_cell.value, status_cell.fill.fgColor.rgb)
        self.assertEqual(set(fills), {"已结构化", "待结构化", "缺失"})
        self.assertEqual(len(set(fills.values())), 3)
        self.assertGreaterEqual(len(quote_sheet.data_validations.dataValidation), 2)
        self.assertGreater(len(quote_sheet.conditional_formatting), 0)

        price_rules_sheet = workbook["报价规则"]
        self.assertEqual(len(price_rules_sheet.data_validations.dataValidation), 1)
        validation = price_rules_sheet.data_validations.dataValidation[0]
        self.assertIn("待确认计算公式", validation.formula1)

        issue_sheet = workbook["数据问题"]
        issue_headers = {cell.value: cell.column for cell in issue_sheet[1]}
        severity_column = issue_headers["严重程度"]
        high_row = next(row for row in issue_sheet.iter_rows(min_row=2) if row[severity_column - 1].value == "高")
        self.assertTrue(high_row[0].fill.fgColor.rgb.endswith("FDE8E7"))


if __name__ == "__main__":
    unittest.main()
