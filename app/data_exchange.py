from __future__ import annotations

import json
import os
import secrets
import shutil
import time
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


DATASETS = [
    ("suppliers", "供应商主数据", "供应商"),
    ("supplier_aliases", "供应商别名", "供应商别名"),
    ("processes", "工艺主数据", "标准工艺"),
    ("process_aliases", "工艺别名", "工艺别名"),
    ("capabilities", "供应商能力", "供应商能力"),
    ("quote_items", "报价明细", "报价项"),
    ("price_rules", "报价规则", "报价规则"),
    ("sku_mappings", "SKU映射", "SKU映射"),
    ("vendor_file_rules", "文件要求", "文件要求"),
    ("issues", "数据问题", "待处理问题"),
]

STATUS_HEADERS = {
    "状态", "映射状态", "尺寸解析状态", "价格解析状态",
    "规则状态", "系统配置状态", "严重程度", "处理状态",
}
TONE_COLORS = {
    "critical": ("FDE8E7", "9F1D14"),
    "pending": ("FFF0C2", "8A5200"),
    "info": ("E8F1FB", "1D4E89"),
    "ready": ("E3F3EC", "0B5C54"),
    "neutral": ("E9EEF3", "52606D"),
}
TONE_PRIORITY = {"critical": 5, "pending": 4, "info": 3, "ready": 2, "neutral": 1}
STATUS_OPTIONS = {
    ("suppliers", "状态"): ["启用待确认", "启用", "停用"],
    ("supplier_aliases", "映射状态"): ["待拆分", "待确认", "已确认"],
    ("process_aliases", "映射状态"): ["待确认", "自动精确匹配", "已确认"],
    ("quote_items", "尺寸解析状态"): ["空", "自动", "自动单边", "待确认", "已确认", "需复核"],
    ("quote_items", "价格解析状态"): ["缺失", "待结构化", "待确认单位", "已结构化", "需复核"],
    ("price_rules", "规则状态"): ["待结构化", "待确认计算公式", "已确认", "停用"],
    ("sku_mappings", "系统配置状态"): ["", "半配置", "已配置", "查不到", "没找到", "首饰类", "配现货 : 仓储部"],
    ("issues", "严重程度"): ["低", "中", "高"],
    ("issues", "处理状态"): ["待处理", "处理中", "已处理", "不处理"],
}


def status_tone(header, value):
    text = str(value or "").strip()
    if header == "严重程度":
        return {"高": "critical", "中": "pending", "低": "info"}.get(text)
    if not text:
        return "neutral" if header in {"系统配置状态", "尺寸解析状态"} else None
    if any(token in text for token in ("缺失", "查不到", "没找到", "异常", "错误", "需复核")):
        return "critical"
    if any(token in text for token in ("待处理", "处理中", "待确认", "待结构化", "待拆分", "半配置")):
        return "pending"
    if text in {"自动", "自动单边", "低"}:
        return "info"
    if text == "启用" or any(token in text for token in ("已确认", "已结构化", "已配置", "自动精确匹配", "已处理")):
        return "ready"
    if text in {"空", "停用", "不处理"}:
        return "neutral"
    return None


def record_tone(record):
    tones = [status_tone(header, record.get(header)) for header in STATUS_HEADERS if header in record]
    tones = [tone for tone in tones if tone]
    return max(tones, key=TONE_PRIORITY.get) if tones else None


def apply_tone(cell, tone, bold=True):
    fill_color, font_color = TONE_COLORS[tone]
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(color=font_color, bold=bold)


def display_width(value):
    text = str(value or "")
    return sum(2 if ord(character) > 127 else 1 for character in text)


def style_table_header(cells, color="17324D"):
    for cell in cells:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_status_controls(sheet, column_letter, options):
    end_row = max(sheet.max_row, 2000)
    target_range = f"{column_letter}2:{column_letter}{end_row}"
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    validation.promptTitle = "选择状态"
    validation.prompt = "请从下拉列表选择，不要输入列表以外的状态。"
    validation.errorTitle = "状态值无效"
    validation.error = "请使用下拉列表中的标准状态。"
    validation.errorStyle = "stop"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(target_range)

    for option in options:
        tone = status_tone(sheet.cell(1, openpyxl.utils.column_index_from_string(column_letter)).value, option)
        if not option or not tone:
            continue
        fill_color, font_color = TONE_COLORS[tone]
        sheet.conditional_formatting.add(
            target_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{option}"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=fill_color),
                font=Font(color=font_color, bold=True),
            ),
        )


def dataset_schemas(reference_source):
    schemas = {}
    for key, _sheet_name, _summary_name in DATASETS:
        rows = reference_source.get(key, [])
        schemas[key] = list(rows[0]) if rows else []
    return schemas


def calculate_summary(source):
    summary = {}
    for key, _sheet_name, summary_name in DATASETS:
        if key in {"supplier_aliases", "process_aliases", "vendor_file_rules"}:
            continue
        summary[summary_name] = len(source.get(key, []))
    return summary


def clean_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def export_xlsx(source):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    info = workbook.create_sheet("导入说明")
    info.sheet_view.showGridLines = False
    info.merge_cells("A1:F1")
    info["A1"] = "报价数据交换与处理说明"
    info["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    info["A1"].fill = PatternFill("solid", fgColor="17324D")
    info["A1"].alignment = Alignment(vertical="center")
    info.row_dimensions[1].height = 36

    info.merge_cells("A2:F2")
    info["A2"] = "后台导出的标准数据文件｜颜色用于标识处理优先级，不改变原始数据值"
    info["A2"].font = Font(color="52606D", italic=True)
    info["A2"].alignment = Alignment(vertical="center")
    info.row_dimensions[2].height = 25
    info["A3"] = "导出时间"
    info.merge_cells("B3:F3")
    info["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    info["A3"].font = Font(bold=True, color="17324D")

    info.merge_cells("A5:F5")
    info["A5"] = "颜色图例与处理顺序"
    info["A5"].font = Font(bold=True, size=13, color="FFFFFF")
    info["A5"].fill = PatternFill("solid", fgColor="0F766E")
    legend = [
        (6, "红色｜必须处理", "缺失、查不到、没找到、异常、需复核或高严重程度；禁止直接用于报价。", "critical"),
        (7, "黄色｜待确认", "待处理、待确认、待结构化、待拆分或半配置；确认后再导入或启用。", "pending"),
        (8, "蓝色｜系统建议", "自动解析、自动单边或低严重程度；可作为辅助信息，仍需抽查。", "info"),
        (9, "绿色｜可直接使用", "已确认、已结构化、已配置或自动精确匹配；可按当前规则查询。", "ready"),
        (10, "灰色｜未提供/不适用", "空白或无尺寸等未提供信息；按业务需要补录，不等同于错误。", "neutral"),
    ]
    for row_number, label, description, tone in legend:
        info[f"A{row_number}"] = label
        apply_tone(info[f"A{row_number}"], tone)
        info.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=6)
        info[f"B{row_number}"] = description
        info[f"B{row_number}"].alignment = Alignment(wrap_text=True, vertical="center")
        info.row_dimensions[row_number].height = 29

    info.merge_cells("A12:F12")
    info["A12"] = "建议处理流程"
    info["A12"].font = Font(bold=True, size=13, color="17324D")
    steps = [
        "1. 先处理红色：补齐缺失值，确认异常与高风险问题。",
        "2. 再处理黄色：确认价格单位、尺寸边界、别名映射和规则公式。",
        "3. 抽查蓝色：核对自动拆分是否与原始文本一致。",
        "4. 绿色数据可以查询使用；最终下单仍应按供应商最新报价复核。",
    ]
    for row_number, step in enumerate(steps, start=13):
        info.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=6)
        info[f"A{row_number}"] = step
        info[f"A{row_number}"].alignment = Alignment(wrap_text=True, vertical="center")
        info.row_dimensions[row_number].height = 24

    info.merge_cells("A18:F18")
    info["A18"] = "导入注意"
    info["A18"].font = Font(bold=True, size=13, color="17324D")
    import_notes = [
        "只修改各数据工作表中的内容，不要修改工作表名、列名或列顺序；导入前系统会预检。",
        "导入为全量替换，确认导入时系统会自动备份当前 JSON 数据。",
        "未审批的尺寸、A3 排版、附加费和 3D 克重/时长规则不会因为导入而自动启用。",
    ]
    for row_number, note in enumerate(import_notes, start=19):
        info.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=6)
        info[f"A{row_number}"] = f"• {note}"
        info[f"A{row_number}"].alignment = Alignment(wrap_text=True, vertical="center")
        info.row_dimensions[row_number].height = 26

    info.merge_cells("A23:F23")
    info["A23"] = "状态怎么修改"
    info["A23"].font = Font(bold=True, size=13, color="FFFFFF")
    info["A23"].fill = PatternFill("solid", fgColor="0F766E")
    status_headers = ["工作表", "状态列", "待处理值", "完成后改为", "修改前必须检查", "系统效果"]
    for column_index, value in enumerate(status_headers, start=1):
        info.cell(24, column_index, value)
    style_table_header(info[24])
    status_guides = [
        ["供应商主数据", "状态", "启用待确认", "启用", "供应商名称、ID、业务范围已复核", "作为可用供应商主数据"],
        ["供应商别名", "映射状态", "待拆分/待确认", "已确认", "别名只对应一个供应商ID", "关联搜索可使用该映射"],
        ["工艺别名", "映射状态", "待确认", "已确认", "工艺原名与标准工艺ID一致", "关联搜索可使用该映射"],
        ["报价明细", "尺寸解析状态", "自动/待确认", "已确认", "毫米宽、高、深与尺寸原文一致", "仅表示源数据已复核；尺寸报价还需后台确认"],
        ["报价明细", "价格解析状态", "缺失/待结构化/待确认单位", "已结构化", "价格上下限为数字，价格类型和计价单位明确", "进入直接报价及历史预估样本"],
        ["报价规则", "规则状态", "待结构化/待确认计算公式", "已确认", "适用工艺、材料、尺寸条件、价格单位完整", "表示规则已审核；未接入的公式仍不自动报价"],
        ["SKU映射", "系统配置状态", "空白/半配置/查不到", "已配置", "工艺、供应商ID、工费字段已补齐", "SKU关联查询可使用该配置"],
        ["数据问题", "处理状态", "待处理/处理中", "已处理", "已回到对应数据表完成修改并复核", "从待处理清单中区分完成项"],
    ]
    for row_number, values in enumerate(status_guides, start=25):
        for column_index, value in enumerate(values, start=1):
            cell = info.cell(row_number, column_index, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        info.row_dimensions[row_number].height = 45
    info.merge_cells("A34:F34")
    info["A34"] = "操作方法：进入对应工作表，在状态列单元格右侧点下拉箭头选择标准状态；不要手工输入其他词。状态改变后，背景颜色会自动更新。"
    info["A34"].alignment = Alignment(wrap_text=True, vertical="center")
    info["A34"].fill = PatternFill("solid", fgColor="E8F1FB")
    info["A34"].font = Font(color="1D4E89", bold=True)
    info.row_dimensions[34].height = 35
    info.merge_cells("A35:F35")
    info["A35"] = "重要：只改状态不会自动补出价格或公式。价格解析状态改为“已结构化”前，必须同时填写价格下限、价格上限和价格类型。"
    info["A35"].alignment = Alignment(wrap_text=True, vertical="center")
    apply_tone(info["A35"], "critical")
    info.row_dimensions[35].height = 35
    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 18
    info.column_dimensions["C"].width = 22
    info.column_dimensions["D"].width = 18
    info.column_dimensions["E"].width = 34
    info.column_dimensions["F"].width = 34

    guide = workbook.create_sheet("规则填写示例")
    guide.sheet_view.showGridLines = False
    guide.merge_cells("A1:H1")
    guide["A1"] = "报价、规则与预估填写示例"
    guide["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="17324D")
    guide["A1"].alignment = Alignment(vertical="center")
    guide.row_dimensions[1].height = 36
    guide.merge_cells("A2:H2")
    guide["A2"] = "本页只用于说明，不参与导入。实际数据请填写到“报价明细”或“报价规则”工作表。"
    guide["A2"].font = Font(color="52606D", italic=True)
    guide.row_dimensions[2].height = 25

    guide.merge_cells("A4:H4")
    guide["A4"] = "一、报价明细怎么写（当前直接报价和历史预估使用这里）"
    guide["A4"].font = Font(bold=True, size=13, color="FFFFFF")
    guide["A4"].fill = PatternFill("solid", fgColor="0F766E")
    quote_example_headers = ["场景", "价格原文", "价格下限", "价格上限", "价格类型", "价格解析状态", "是否进入预估", "填写说明"]
    for column_index, value in enumerate(quote_example_headers, start=1):
        guide.cell(5, column_index, value)
    style_table_header(guide[5])
    quote_examples = [
        ["固定单价", "3元/个", 3, 3, "固定单价", "已结构化", "是", "表示固定单价3元/个，不是总价；上下限填同一数字。"],
        ["价格区间", "5-8元/个", 5, 8, "区间价", "已结构化", "是", "下限不得大于上限；历史预估当前以价格下限作为样本。"],
        ["单位待确认", "3元（未注明按个或按版）", "", "", "原文待确认", "待确认单位", "否", "先向供应商确认计价单位，再补数字并改为已结构化。"],
        ["无法拆分", "按数量、颜色另算", "", "", "复杂规则", "待结构化", "否", "不要猜价格；保留原文，转到报价规则表拆分。"],
    ]
    for row_number, values in enumerate(quote_examples, start=6):
        for column_index, value in enumerate(values, start=1):
            guide.cell(row_number, column_index, value).alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[row_number].height = 42

    guide.merge_cells("A11:H11")
    guide["A11"] = "二、报价规则怎么写"
    guide["A11"].font = Font(bold=True, size=13, color="FFFFFF")
    guide["A11"].fill = PatternFill("solid", fgColor="0F766E")
    rule_example_headers = ["规则场景", "规则类型", "关联工艺原名", "材料", "尺寸条件原文", "价格字段写法", "规则状态", "当前系统处理"]
    for column_index, value in enumerate(rule_example_headers, start=1):
        guide.cell(12, column_index, value)
    style_table_header(guide[12])
    rule_examples = [
        ["固定价", "固定价", "水晶标", "PVC", "宽<=200mm；高<=300mm", "下限=3；上限=3；单位=元/个", "已确认", "建议同步写入报价明细；仅写规则表暂不进入自动预估。"],
        ["精确尺寸价", "尺寸固定价", "贺卡打印", "卡纸", "宽=95mm；高=70mm", "下限=1；上限=1；单位=元/个", "已确认", "需有同条件报价明细且后台尺寸边界已确认后才匹配。"],
        ["尺寸范围价", "尺寸分档", "印刷", "卡纸", "宽60-90mm；高80-90mm", "下限=1；上限=1.5；单位=元/个", "已确认", "目前不做尺寸插值；范围规则只作为人工审核依据。"],
        ["附加费", "附加费", "通用", "", "加袋", "下限=0.2；上限=0.2；单位=元/个", "已确认", "当前不会自动叠加到报价，需人工加入最终报价。"],
        ["A3排版", "A3排版", "A3印刷", "", "成品297*420mm；排版数量=人工确认", "写清按版或按个的价格单位", "待确认计算公式", "当前不自动计算排版数，必须人工核价。"],
        ["3D打印", "3D克重时长", "3D打印", "PLA", "长165mm；宽115mm；厚60mm", "克重单价、时长单价、损耗需分别写清", "待确认计算公式", "当前不自动计算克重、支撑损耗和打印时长。"],
        ["历史预估", "无需新增公式", "填写具体工艺", "填写具体材料", "目标尺寸在后台已确认边界内", "使用已结构化报价明细", "不适用", "系统按同工艺/材料/供应商历史固定价筛选并给出中位参考。"],
    ]
    for row_number, values in enumerate(rule_examples, start=13):
        for column_index, value in enumerate(values, start=1):
            guide.cell(row_number, column_index, value).alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[row_number].height = 55

    guide.merge_cells("A22:H22")
    guide["A22"] = "三、尺寸条件建议写法"
    guide["A22"].font = Font(bold=True, size=13, color="FFFFFF")
    guide["A22"].fill = PatternFill("solid", fgColor="0F766E")
    size_examples = [
        ["类型", "标准写法", "示例含义", "注意事项"],
        ["精确矩形", "宽=95mm；高=70mm", "只适用于95×70mm，可旋转匹配", "不要省略单位"],
        ["上限矩形", "宽<=200mm；高<=300mm", "目标尺寸不得超过200×300mm", "“以内”统一拆成上限"],
        ["范围矩形", "宽60-90mm；高80-90mm", "宽高都必须落在各自范围内", "不要只写一个总面积"],
        ["标准纸张", "A3；成品297*420mm", "A3成品尺寸", "排版数量、出血和间距另行确认"],
        ["三维尺寸", "长165mm；宽115mm；厚60mm", "三维成品尺寸", "尺寸完整也不代表3D价格公式已确认"],
        ["直径上限", "直径<=45mm", "圆形尺寸不超过45mm", "不要写成长宽矩形"],
    ]
    for row_offset, values in enumerate(size_examples, start=23):
        for column_index, value in enumerate(values, start=1):
            guide.cell(row_offset, column_index, value)
            guide.cell(row_offset, column_index).alignment = Alignment(wrap_text=True, vertical="top")
        if row_offset == 23:
            style_table_header(guide[row_offset][:4])
        else:
            guide.row_dimensions[row_offset].height = 38

    guide.merge_cells("A31:H31")
    guide["A31"] = "四、固定单价怎么计价"
    guide["A31"].font = Font(bold=True, size=13, color="FFFFFF")
    guide["A31"].fill = PatternFill("solid", fgColor="0F766E")
    pricing_headers = ["项目", "示例值", "单位", "是否必填", "计算作用", "当前系统处理", "风险提示", "业务解释"]
    for column_index, value in enumerate(pricing_headers, start=1):
        guide.cell(32, column_index, value)
    style_table_header(guide[32])
    pricing_examples = [
        ["单价下限", 3, "元/个", "是", "固定单价计算值", "已结构化后作为历史价格样本", "不要填单位文字", "每个最低3元"],
        ["单价上限", 3, "元/个", "是", "与下限相同表示固定价", "保存为价格范围", "上限不得小于下限", "每个最高3元"],
        ["需求数量", 100, "个", "计算总价时必填", "单价×数量", "当前预估页未接入数量", "需确认数量分档", "本次询价100个"],
        ["固定附加费", 0, "元", "无则填0", "打样、开机、加急等固定费用", "当前不自动叠加", "与按个附加费分开", "本示例无附加费"],
        ["最低起订金额", 0, "元", "无则填0", "总价不得低于此金额", "当前不自动判断", "不等于起订数量", "本示例无最低金额"],
        ["理论总价", "=MAX(B33*B35+B36,B37)", "元", "系统计算", "MAX(单价×数量+固定附加费,最低起订金额)", "本页公式仅用于演示", "不含运费、税费和按个附加费", "3×100+0=300元"],
    ]
    for row_number, values in enumerate(pricing_examples, start=33):
        for column_index, value in enumerate(values, start=1):
            cell = guide.cell(row_number, column_index)
            if row_number == 38 and column_index == 2:
                cell.value = value
                cell.number_format = '¥#,##0.00'
            else:
                cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[row_number].height = 48
    for row_number in range(33, 38):
        guide.cell(row_number, 2).number_format = '#,##0.00'
    guide.merge_cells("A40:H40")
    guide["A40"] = "当前系统边界：预估只读取“报价明细”中价格解析状态=已结构化的价格下限作为单价历史样本；不自动读取报价规则表、不自动乘数量、不叠加附加费。"
    guide["A40"].alignment = Alignment(wrap_text=True, vertical="center")
    apply_tone(guide["A40"], "pending")
    guide.row_dimensions[40].height = 38
    guide.freeze_panes = "A5"
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 24
    guide.column_dimensions["C"].width = 22
    guide.column_dimensions["D"].width = 20
    guide.column_dimensions["E"].width = 34
    guide.column_dimensions["F"].width = 32
    guide.column_dimensions["G"].width = 22
    guide.column_dimensions["H"].width = 42

    summary_sheet = workbook.create_sheet("数据总览")
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.append(["数据分类", "记录数", "红色必须处理", "黄色待确认", "蓝色系统建议", "绿色可使用", "灰色空白"])
    for key, sheet_name, _summary_name in DATASETS:
        counts = {tone: 0 for tone in TONE_COLORS}
        for record in source.get(key, []):
            tone = record_tone(record)
            if tone:
                counts[tone] += 1
        summary_sheet.append([
            sheet_name, len(source.get(key, [])), counts["critical"], counts["pending"],
            counts["info"], counts["ready"], counts["neutral"],
        ])
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = summary_sheet.dimensions
    summary_sheet.row_dimensions[1].height = 28
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in summary_sheet.iter_rows(min_row=2, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
    for column_index, tone in enumerate((None, "critical", "pending", "info", "ready", "neutral"), start=2):
        if tone:
            fill_color, font_color = TONE_COLORS[tone]
            for cell in summary_sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=summary_sheet.max_row).__next__():
                if cell.value:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.font = Font(color=font_color, bold=True)
    summary_sheet.column_dimensions["A"].width = 22
    for column in "BCDEFG":
        summary_sheet.column_dimensions[column].width = 17

    schemas = dataset_schemas(source)
    for key, sheet_name, _summary_name in DATASETS:
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        headers = schemas[key]
        sheet.append(headers)
        for record in source.get(key, []):
            sheet.append([record.get(header, "") for header in headers])
            row_cells = sheet[sheet.max_row]
            if key == "issues":
                issue_tone = status_tone("严重程度", record.get("严重程度"))
                if issue_tone:
                    fill_color, font_color = TONE_COLORS[issue_tone]
                    for cell in row_cells:
                        cell.fill = PatternFill("solid", fgColor=fill_color)
                        cell.font = Font(color=font_color)
            for index, cell in enumerate(row_cells):
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.data_type = "s"
                header = headers[index]
                if header in STATUS_HEADERS:
                    tone = status_tone(header, cell.value)
                    if tone:
                        apply_tone(cell, tone)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 28
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for index, header in enumerate(headers, start=1):
            values = [header, *[record.get(header, "") for record in source.get(key, [])[:200]]]
            width = min(max(12, max(display_width(value) for value in values) + 2), 42)
            column_letter = openpyxl.utils.get_column_letter(index)
            sheet.column_dimensions[column_letter].width = width
            status_options = STATUS_OPTIONS.get((key, header))
            if status_options:
                add_status_controls(sheet, column_letter, status_options)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_json_import(content):
    try:
        source = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise ValueError("JSON 文件无法解析") from error
    if not isinstance(source, dict):
        raise ValueError("JSON 根节点必须是对象")
    return source


def parse_xlsx_import(content, reference_source):
    try:
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as error:
        raise ValueError("Excel 文件无法解析") from error
    schemas = dataset_schemas(reference_source)
    source = {}
    for key, sheet_name, _summary_name in DATASETS:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"缺少工作表：{sheet_name}")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=False)
        try:
            header_cells = next(rows)
        except StopIteration as error:
            raise ValueError(f"工作表 {sheet_name} 为空") from error
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        expected = schemas[key]
        if headers[: len(expected)] != expected:
            raise ValueError(f"工作表 {sheet_name} 列名或顺序已改变")
        records = []
        for row_number, cells in enumerate(rows, start=2):
            relevant = cells[: len(expected)]
            if any(cell.data_type == "f" for cell in relevant):
                raise ValueError(f"工作表 {sheet_name} 第 {row_number} 行含公式，请改为固定值")
            values = [clean_excel_value(cell.value) for cell in relevant]
            if not any(value not in ("", None) for value in values):
                continue
            records.append(dict(zip(expected, values)))
        source[key] = records
    return source


def validate_source(source, reference_source):
    errors = []
    schemas = dataset_schemas(reference_source)
    counts = {}
    for key, sheet_name, _summary_name in DATASETS:
        rows = source.get(key)
        if not isinstance(rows, list):
            errors.append(f"{sheet_name}：缺失或不是数组")
            continue
        counts[sheet_name] = len(rows)
        required_headers = schemas[key]
        id_header = next((header for header in required_headers if header.endswith("ID")), None)
        seen_ids = set()
        for index, record in enumerate(rows, start=2):
            if not isinstance(record, dict):
                errors.append(f"{sheet_name}第 {index} 行不是有效记录")
                continue
            missing = [header for header in required_headers if header not in record]
            if missing:
                errors.append(f"{sheet_name}第 {index} 行缺少列：{', '.join(missing[:3])}")
            if id_header:
                value = str(record.get(id_header, "")).strip()
                if not value:
                    errors.append(f"{sheet_name}第 {index} 行缺少 {id_header}")
                elif value in seen_ids:
                    errors.append(f"{sheet_name}：{id_header} {value} 重复")
                seen_ids.add(value)
            if len(errors) >= 50:
                break
        if len(errors) >= 50:
            break
    source["summary"] = calculate_summary(source)
    return counts, errors


def parse_and_validate(content, file_format, reference_source):
    if file_format == "json":
        source = parse_json_import(content)
    elif file_format == "xlsx":
        source = parse_xlsx_import(content, reference_source)
    else:
        raise ValueError("仅支持 .xlsx 和 .json 文件")
    counts, errors = validate_source(source, reference_source)
    return source, counts, errors


def stage_import(staging_dir, source, user_id):
    staging_dir.mkdir(parents=True, exist_ok=True)
    token = secrets.token_urlsafe(24)
    payload = {
        "userId": user_id,
        "expiresAt": int(time.time()) + 15 * 60,
        "source": source,
    }
    target = staging_dir / f"{token}.json"
    target.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    os.chmod(target, 0o600)
    return token


def commit_staged_import(staging_dir, token, user_id, source_path, backup_dir, rebuild_callback):
    if not token or not all(character.isalnum() or character in "-_" for character in token):
        raise ValueError("无效的导入凭证")
    staged_path = staging_dir / f"{token}.json"
    if not staged_path.is_file():
        raise ValueError("导入预检已过期，请重新上传")
    payload = json.loads(staged_path.read_text(encoding="utf-8"))
    if payload.get("userId") != user_id or payload.get("expiresAt", 0) < int(time.time()):
        staged_path.unlink(missing_ok=True)
        raise ValueError("导入预检已过期，请重新上传")

    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"master_data_source_{timestamp}.json"
    if source_path.exists():
        shutil.copy2(source_path, backup_path)

    temporary = source_path.with_suffix(".json.importing")
    temporary.write_text(json.dumps(payload["source"], ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, source_path)
    try:
        summary = rebuild_callback()
    except Exception:
        if backup_path.exists():
            shutil.copy2(backup_path, source_path)
            rebuild_callback()
        raise
    staged_path.unlink(missing_ok=True)
    return summary, backup_path
