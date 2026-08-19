from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILES = {
    "capability": ROOT / "副本定制工艺及供应商时效2026.7.1.xlsx",
    "quotes": ROOT / "副本供应商工艺沉淀 2026.7.16.xlsx",
    "dictionary": ROOT / "副本定制工艺共享.xlsx",
}


def text(value):
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").strip()


def actual_rows(sheet, max_rows=1200, max_cols=50):
    rows = []
    for values in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row, max_rows),
        max_col=min(sheet.max_column, max_cols),
        values_only=True,
    ):
        rows.append(list(values))
    last = 0
    for index, row in enumerate(rows, start=1):
        if any(value not in (None, "") for value in row):
            last = index
    return rows[:last]


def row_has_content(row):
    return any(value not in (None, "") for value in row)


def stable_id(prefix, value):
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8].upper()
    return f"{prefix}-{digest}"


def normalize_process(value):
    return re.sub(r"\s+", "", text(value)).replace("（", "(").replace("）", ")")


def parse_simple_dimension(value):
    raw = text(value)
    if not raw:
        return "", "", "", "空"
    cleaned = raw.replace("×", "*").replace("x", "*").replace("X", "*").replace("#", "*")
    number = r"(\d+(?:\.\d+)?)"
    pattern = re.compile(
        rf"^\s*(?:尺寸[:：]?\s*)?{number}\s*\*\s*{number}(?:\s*\*\s*{number})?\s*(mm|毫米|cm|CM|厘米)\s*$"
    )
    match = pattern.match(cleaned)
    if not match:
        one = re.match(rf"^\s*(?:尺寸[:：]?\s*)?{number}\s*(mm|毫米|cm|CM|厘米)\s*$", cleaned)
        if not one:
            return "", "", "", "待确认"
        value_mm = float(one.group(1)) * (10 if one.group(2).lower() in ("cm", "厘米") else 1)
        return value_mm, "", "", "自动单边"
    unit = match.group(4).lower()
    factor = 10 if unit in ("cm", "厘米") else 1
    values = [float(match.group(1)) * factor, float(match.group(2)) * factor]
    if match.group(3):
        values.append(float(match.group(3)) * factor)
    while len(values) < 3:
        values.append("")
    return values[0], values[1], values[2], "自动"


def classify_price(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value), float(value), "固定单价", "已结构化"
    raw = text(value)
    if not raw:
        return "", "", "", "缺失"
    single = re.fullmatch(r"\s*(\d+(?:\.\d+)?)\s*元?\s*", raw)
    if single:
        price = float(single.group(1))
        return price, price, "文本单价", "待确认单位"
    return "", "", "", "待结构化"


def main(output_path):
    books = {
        name: openpyxl.load_workbook(path, read_only=True, data_only=False)
        for name, path in SOURCE_FILES.items()
    }

    suppliers = OrderedDict()
    processes = OrderedDict()
    capabilities = []
    quote_items = []
    price_rules = []
    sku_mappings = []
    aliases = OrderedDict()
    issues = []

    def register_supplier(name, source_tag, supplier_type="供应商"):
        name = text(name)
        if not name:
            return ""
        if name not in suppliers:
            suppliers[name] = {
                "供应商ID": stable_id("SUP", name),
                "供应商名称": name,
                "供应商类型": supplier_type,
                "来源": set(),
                "状态": "启用待确认",
                "备注": "",
            }
        suppliers[name]["来源"].add(source_tag)
        if "/" in name and suppliers[name]["供应商类型"] == "供应商":
            suppliers[name]["供应商类型"] = "供应商组合"
        return suppliers[name]["供应商ID"]

    def supplier_id(name):
        return suppliers.get(text(name), {}).get("供应商ID", "")

    def add_issue(issue_type, severity, source, detail, suggested_action):
        issues.append(
            {
                "问题ID": f"ISS-{len(issues) + 1:04d}",
                "问题类型": issue_type,
                "严重程度": severity,
                "来源": source,
                "问题说明": detail,
                "建议处理": suggested_action,
                "处理状态": "待处理",
            }
        )

    # Standard process dictionary and vendor file requirements.
    dictionary = books["dictionary"]
    process_sheet = dictionary["Sheet1"]
    category = ""
    for source_row, row in enumerate(actual_rows(process_sheet)[2:], start=3):
        if not row_has_content(row):
            continue
        category = text(row[0]) or category
        name = text(row[1])
        if not name:
            continue
        if name not in processes:
            processes[name] = {
                "工艺ID": stable_id("PROC", name),
                "一级分类": category,
                "标准工艺名称": name,
                "通用文件要求": text(row[2]),
                "适用材料": text(row[3]),
                "通用备注": text(row[4]),
                "来源": f"副本定制工艺共享.xlsx!Sheet1!{source_row}",
            }

    vendor_file_rules = []
    file_sheet = dictionary["各供应商定制文件要求"]
    current_supplier = ""
    for source_row, row in enumerate(actual_rows(file_sheet)[1:], start=2):
        if not row_has_content(row):
            continue
        current_supplier = text(row[0]) or current_supplier
        if current_supplier:
            register_supplier(current_supplier, "定制文件要求")
        vendor_file_rules.append(
            {
                "供应商原名": current_supplier,
                "工艺原名": text(row[1]),
                "文件格式": text(row[2]),
                "分辨率": text(row[3]),
                "像素": text(row[4]),
                "尺寸要求": text(row[5]),
                "来源": f"副本定制工艺共享.xlsx!各供应商定制文件要求!{source_row}",
            }
        )

    # Supplier capability and delivery-time records.
    capability_book = books["capability"]
    capability_sheet = capability_book["供应商维度"]
    current_supplier = ""
    raw_process_counts = OrderedDict()
    rows = actual_rows(capability_sheet)
    for source_row, row in enumerate(rows[1:], start=2):
        if not row_has_content(row):
            continue
        current_supplier = text(row[0]) or current_supplier
        supplier_type = "供应商组合" if "/" in current_supplier else "供应商"
        sup_id = register_supplier(current_supplier, "供应商工艺时效", supplier_type)
        secondary = text(row[2])
        if secondary:
            raw_process_counts[secondary] = raw_process_counts.get(secondary, 0) + 1
        source = f"副本定制工艺及供应商时效2026.7.1.xlsx!供应商维度!{source_row}"
        capabilities.append(
            {
                "能力ID": f"CAP-{len(capabilities) + 1:04d}",
                "供应商ID": sup_id,
                "供应商原名": current_supplier,
                "一级工艺原名": text(row[1]),
                "二级工艺原名": secondary,
                "标准工艺ID": "",
                "生产时效_天": row[4] if len(row) > 4 else "",
                "物流时效_天": row[5] if len(row) > 5 else "",
                "跟单复核": text(row[6]) if len(row) > 6 else "",
                "备注": text(row[7]) if len(row) > 7 else "",
                "来源": source,
            }
        )
        if row[4] in (None, ""):
            add_issue("生产时效缺失", "高", source, f"{current_supplier} / {secondary}", "补充更新后生产时效")
        if row[6] in (None, ""):
            add_issue("跟单复核缺失", "低", source, f"{current_supplier} / {secondary}", "补充确认状态或标记不适用")

    # Supplier-specific quote sheets.
    quote_book = books["quotes"]
    supplier_sheets = [sheet for sheet in quote_book.worksheets if sheet.title not in {"共用SKU", "208个SKU", "WpsReserved_CellImgList"}]
    for sheet in supplier_sheets:
        rows = actual_rows(sheet)
        if not rows:
            continue
        sheet_supplier = sheet.title
        register_supplier(sheet_supplier, "供应商报价沉淀")
        source_base = f"副本供应商工艺沉淀 2026.7.16.xlsx!{sheet.title}"
        base_rule = text(rows[0][0]) if rows[0] else ""
        if base_rule:
            price_rules.append(
                {
                    "规则ID": f"RULE-{len(price_rules) + 1:04d}",
                    "供应商ID": supplier_id(sheet_supplier),
                    "供应商原名": sheet_supplier,
                    "规则类型": "供应商基础规则",
                    "关联工艺原名": "",
                    "材料": "",
                    "尺寸条件原文": "",
                    "价格下限": "",
                    "价格上限": "",
                    "价格单位": "",
                    "规则原文": base_rule,
                    "规则状态": "待结构化",
                    "来源": f"{source_base}!1",
                }
            )

        if sheet.title == "弘景天":
            for source_row, row in enumerate(rows[2:], start=3):
                if not row_has_content(row):
                    continue
                material = text(row[0])
                color = text(row[1])
                gram = text(row[2])
                hour = text(row[3])
                maximum = text(row[4])
                price_rules.append(
                    {
                        "规则ID": f"RULE-{len(price_rules) + 1:04d}",
                        "供应商ID": supplier_id(sheet_supplier),
                        "供应商原名": sheet_supplier,
                        "规则类型": "3D打印_克重时长",
                        "关联工艺原名": "3D打印",
                        "材料": material,
                        "尺寸条件原文": maximum,
                        "价格下限": "",
                        "价格上限": "",
                        "价格单位": "元/g + 元/h",
                        "规则原文": f"颜色:{color}; 克重:{gram}; 时长:{hour}; 最大尺寸:{maximum}; 备注:{text(row[6])}",
                        "规则状态": "待确认计算公式",
                        "来源": f"{source_base}!{source_row}",
                    }
                )
            continue

        if len(rows) < 2:
            continue
        headers = [text(value) for value in rows[1]]

        def values_for(row, header):
            return [row[index] for index, value in enumerate(headers) if value == header and index < len(row)]

        def first_value(row, *header_names):
            for header_name in header_names:
                for value in values_for(row, header_name):
                    if value not in (None, ""):
                        return value
            return ""

        data_count = 0
        for source_row, row in enumerate(rows[2:], start=3):
            if not row_has_content(row):
                continue
            # Rows containing only image formulas are not usable quote records.
            non_formula_values = [value for value in row if not (isinstance(value, str) and "DISPIMG" in value)]
            if not row_has_content(non_formula_values):
                continue
            data_count += 1
            supplier_raw = text(first_value(row, "供应商")) or sheet_supplier
            sup_id = register_supplier(supplier_raw, "供应商报价沉淀")
            process_raw = text(first_value(row, "工艺"))
            if process_raw:
                raw_process_counts[process_raw] = raw_process_counts.get(process_raw, 0) + 1
            product_size = text(first_value(row, "产品尺寸", "名称/尺寸"))
            custom_size = text(first_value(row, "定制尺寸"))
            width, height, depth, dimension_status = parse_simple_dimension(custom_size)
            price_raw = first_value(row, "参考价格")
            price_min, price_max, price_type, price_status = classify_price(price_raw)
            source = f"{source_base}!{source_row}"
            quote_items.append(
                {
                    "报价项ID": f"QUOTE-{len(quote_items) + 1:04d}",
                    "供应商ID": sup_id,
                    "供应商原名": supplier_raw,
                    "SKU": text(first_value(row, "SKU")),
                    "共用SKU原文": text(first_value(row, "共用SKU")),
                    "工艺原名": process_raw,
                    "标准工艺ID": "",
                    "材质": text(first_value(row, "材质")),
                    "产品尺寸原文": product_size,
                    "定制尺寸原文": custom_size,
                    "定制宽_mm": width,
                    "定制高_mm": height,
                    "定制深_mm": depth,
                    "尺寸解析状态": dimension_status,
                    "生产时效_天": first_value(row, "生产参考时效"),
                    "物流时效_天": first_value(row, "物流参考时效"),
                    "文件要求": text(first_value(row, "文件要求", "文件要求（先不管）")),
                    "价格原文": text(price_raw),
                    "价格下限": price_min,
                    "价格上限": price_max,
                    "价格类型": price_type,
                    "价格解析状态": price_status,
                    "注意事项": text(first_value(row, "注意事项")),
                    "来源": source,
                }
            )
            if price_status != "已结构化":
                add_issue("报价待结构化", "高" if price_status == "缺失" else "中", source, f"原始报价: {text(price_raw) or '空'}", "确认价格单位、分档或计算公式")
            if custom_size and dimension_status == "待确认":
                add_issue("尺寸待结构化", "中", source, f"定制尺寸: {custom_size}", "拆分为宽、高、深和单位，或配置尺寸规则")

        if data_count == 0 and sheet.title in {"咏帆", "新触觉", "杰洲"}:
            add_issue("空供应商模板", "低", f"{source_base}!2", "只有表头，无可查询报价记录", "补录报价，或在系统中停用该供应商")

    # SKU mapping and configuration status.
    config_status = {}
    config_sheet = quote_book["208个SKU"]
    for source_row, row in enumerate(actual_rows(config_sheet)[1:], start=2):
        if row_has_content(row) and text(row[0]):
            config_status[text(row[0])] = text(row[1])

    mapping_sheet = quote_book["共用SKU"]
    for source_row, row in enumerate(actual_rows(mapping_sheet)[1:], start=2):
        if not row_has_content(row) or not text(row[0]):
            continue
        supplier_1 = text(row[2]) if len(row) > 2 else ""
        supplier_2 = text(row[5]) if len(row) > 5 else ""
        if supplier_1:
            register_supplier(supplier_1, "共用SKU映射")
        if supplier_2:
            register_supplier(supplier_2, "共用SKU映射")
        sku_mappings.append(
            {
                "映射ID": f"SKU-{len(sku_mappings) + 1:04d}",
                "SKU": text(row[0]),
                "系统配置状态": config_status.get(text(row[0]), ""),
                "工艺1原名": text(row[1]) if len(row) > 1 else "",
                "供应商1原名": supplier_1,
                "供应商1_ID": supplier_id(supplier_1),
                "工费1": row[3] if len(row) > 3 else "",
                "工艺2原名": text(row[4]) if len(row) > 4 else "",
                "供应商2原名": supplier_2,
                "供应商2_ID": supplier_id(supplier_2),
                "工费2": row[6] if len(row) > 6 else "",
                "来源": f"副本供应商工艺沉淀 2026.7.16.xlsx!共用SKU!{source_row}",
            }
        )

    # Map exact process aliases only. Complex combinations remain for review.
    standard_by_normalized = {normalize_process(name): item["工艺ID"] for name, item in processes.items()}
    for raw_name, source_count in raw_process_counts.items():
        normalized = normalize_process(raw_name)
        exact_process_id = standard_by_normalized.get(normalized, "")
        aliases[raw_name] = {
            "别名ID": stable_id("PAL", raw_name),
            "工艺原名": raw_name,
            "标准工艺ID": exact_process_id,
            "映射状态": "自动精确匹配" if exact_process_id else "待确认",
            "来源记录数": source_count,
            "处理说明": "仅按名称完全一致自动映射；组合工艺和近义词需人工确认。",
        }

    for item in capabilities:
        item["标准工艺ID"] = aliases.get(item["二级工艺原名"], {}).get("标准工艺ID", "")
    for item in quote_items:
        item["标准工艺ID"] = aliases.get(item["工艺原名"], {}).get("标准工艺ID", "")
    for item in vendor_file_rules:
        raw = item["工艺原名"]
        if raw and raw not in aliases:
            exact_process_id = standard_by_normalized.get(normalize_process(raw), "")
            aliases[raw] = {
                "别名ID": stable_id("PAL", raw),
                "工艺原名": raw,
                "标准工艺ID": exact_process_id,
                "映射状态": "自动精确匹配" if exact_process_id else "待确认",
                "来源记录数": 1,
                "处理说明": "来自供应商文件要求。",
            }

    for alias in aliases.values():
        if alias["映射状态"] == "待确认":
            add_issue("工艺别名待映射", "中", "工艺别名", alias["工艺原名"], "指定标准工艺，或新增标准工艺")

    supplier_aliases = []
    for supplier in suppliers.values():
        name = supplier["供应商名称"]
        supplier_aliases.append(
            {
                "别名ID": stable_id("SAL", name),
                "供应商ID": supplier["供应商ID"],
                "别名": name,
                "别名类型": "原始名称" if supplier["供应商类型"] != "供应商组合" else "组合名称",
                "映射状态": "待拆分" if supplier["供应商类型"] == "供应商组合" else "已确认",
                "来源": "；".join(sorted(supplier["来源"])),
                "备注": "",
            }
        )
        if supplier["供应商类型"] == "供应商组合":
            add_issue("供应商组合待拆分", "中", "供应商主数据", name, "确认组合内供应商及其对应报价、时效")

    supplier_rows = []
    for supplier in suppliers.values():
        supplier_rows.append(
            {
                **{key: value for key, value in supplier.items() if key != "来源"},
                "来源": "；".join(sorted(supplier["来源"])),
            }
        )

    # Excel error values from the source must not become queryable values or
    # formula-like errors in the cleaned workbook.
    source_collections = [capabilities, quote_items, price_rules, sku_mappings, vendor_file_rules]
    error_pattern = re.compile(r"^#(?:N/A|REF!|DIV/0!|VALUE!|NAME\\?|NUM!|NULL!)$")
    for collection in source_collections:
        for record in collection:
            for field, value in list(record.items()):
                if isinstance(value, str) and error_pattern.match(value.strip()):
                    add_issue(
                        "源数据错误值",
                        "中",
                        text(record.get("来源")) or "原始来源未记录",
                        f"{field}: 原始错误值（{value[1:]}）",
                        "确认实际值，或保留为空并标记不可报价",
                    )
                    record[field] = f"原始错误值（{value[1:]}）"

    data = {
        "summary": {
            "供应商": len(supplier_rows),
            "标准工艺": len(processes),
            "供应商能力": len(capabilities),
            "报价项": len(quote_items),
            "报价规则": len(price_rules),
            "SKU映射": len(sku_mappings),
            "待处理问题": len(issues),
        },
        "suppliers": supplier_rows,
        "supplier_aliases": supplier_aliases,
        "processes": list(processes.values()),
        "process_aliases": list(aliases.values()),
        "capabilities": capabilities,
        "quote_items": quote_items,
        "price_rules": price_rules,
        "sku_mappings": sku_mappings,
        "vendor_file_rules": vendor_file_rules,
        "issues": issues,
    }
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    Path(output_path).write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "outputs" / "master_data_source.json"
    main(target)
